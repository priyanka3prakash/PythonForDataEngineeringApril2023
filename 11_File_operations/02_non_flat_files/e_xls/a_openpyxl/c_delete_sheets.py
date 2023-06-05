#!/usr/bin/python3
import openpyxl


def create_worksheets(path):
    workbook = openpyxl.Workbook()
    workbook.create_sheet()

    # Insert a worksheet
    workbook.create_sheet(index=1, title="Second sheet")
    print(workbook.sheetnames)

    del workbook["Second sheet"]
    print(workbook.sheetnames)

    workbook.save(path)


def remove_worksheets(path):
    workbook = openpyxl.Workbook()
    sheet1 = workbook.create_sheet()

    # Insert a worksheet
    workbook.create_sheet(index=1, title="Second sheet")
    print(workbook.sheetnames)
    workbook.remove(sheet1)
    print(workbook.sheetnames)
    workbook.save(path)


if __name__ == "__main__":
    remove_worksheets("d_remove_sheets.xlsx")
    create_worksheets("c_delete_sheets.xlsx")
